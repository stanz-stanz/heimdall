"""CVR Excel reader with column header validation."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from loguru import logger

from src.prospecting.config import (
    COL_ADDRESS,
    COL_AD_PROTECTED,
    COL_CITY,
    COL_COMPANY_FORM,
    COL_CVR,
    COL_EMAIL,
    COL_INDUSTRY,
    COL_NAME,
    COL_PHONE,
    COL_POSTCODE,
)

# Expected substrings in column headers (case-insensitive) for validation
EXPECTED_HEADERS: dict[int, str] = {
    COL_CVR: "cvr",
    COL_NAME: "navn",
    COL_ADDRESS: "adresse",
    COL_POSTCODE: "postnr",
    COL_CITY: "by",
    COL_COMPANY_FORM: "virksomhedsform",
    COL_INDUSTRY: "branche",
    COL_PHONE: "telefon",
    COL_EMAIL: "email",
    COL_AD_PROTECTED: "reklame",
}


class HeaderMismatchError(Exception):
    """Raised when Excel column headers don't match expected layout."""


def validate_headers(ws) -> None:
    """Validate that the first row contains expected column headers.

    Raises HeaderMismatchError with details if validation fails.
    """
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    mismatches = []

    for col_idx, expected_substr in EXPECTED_HEADERS.items():
        if col_idx >= len(header_row) or header_row[col_idx] is None:
            mismatches.append(
                f"  col {col_idx}: expected '{expected_substr}', got: (missing)"
            )
            continue
        actual = str(header_row[col_idx]).lower()
        if expected_substr not in actual:
            mismatches.append(
                f"  col {col_idx}: expected '{expected_substr}', got: '{header_row[col_idx]}'"
            )

    if mismatches:
        detail = "\n".join(mismatches)
        raise HeaderMismatchError(
            f"Excel column headers do not match expected layout:\n{detail}\n"
            f"Check that config column indices match the actual Excel file."
        )


def _parse_industry(raw: str) -> tuple[str, str]:
    """Split '468600 Engroshandel med ...' into (code, danish_name)."""
    if not raw:
        return "", ""
    parts = raw.strip().split(" ", 1)
    code = parts[0] if parts else ""
    name = parts[1] if len(parts) > 1 else ""
    return code, name


def read_cvr_excel(path: Path) -> list[dict]:
    """Read the CVR Excel export and return a list of row dicts.

    Validates column headers before processing. Raises HeaderMismatchError
    if the layout doesn't match expectations.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    validate_headers(ws)

    rows = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[COL_CVR]:
            continue

        industry_code, industry_name_da = _parse_industry(str(row[COL_INDUSTRY] or ""))

        rows.append({
            "cvr": str(row[COL_CVR]).strip(),
            "name": str(row[COL_NAME] or "").strip(),
            "address": str(row[COL_ADDRESS] or "").strip(),
            "postcode": str(row[COL_POSTCODE] or "").strip(),
            "city": str(row[COL_CITY] or "").strip(),
            "company_form": str(row[COL_COMPANY_FORM] or "").strip(),
            "industry_code": industry_code,
            "industry_name_da": industry_name_da,
            "phone": str(row[COL_PHONE] or "").strip(),
            "email": str(row[COL_EMAIL] or "").strip().lower() if row[COL_EMAIL] else "",
            "ad_protected": 1 if str(row[COL_AD_PROTECTED] or "").strip().lower() == "ja" else 0,
            "source_file": path.name,
            "source_row": row_num,
        })

    wb.close()
    logger.info("Read {} companies from {}", len(rows), path.name)
    return rows
